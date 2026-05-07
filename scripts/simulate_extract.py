#!/usr/bin/env python3
"""
Rust 추출기와 등가의 휴리스틱(+PyMuPDF 또는 pdfplumber)을 적용해 data.json
시제품을 만든다. cargo가 없는 환경에서 결과 분포를 빠르게 점검하기 위한 도구.

사용:
  python3 scripts/simulate_extract.py data site/public/data.json
"""
from __future__ import annotations
import argparse, hashlib, json, os, re, sys, time
from pathlib import Path

try:
    import fitz  # PyMuPDF
    _HAS_FITZ = True
except ImportError:
    _HAS_FITZ = False
try:
    import pdfplumber  # 폴백
except ImportError:
    pdfplumber = None
if not _HAS_FITZ and pdfplumber is None:
    print("pip install pymupdf (또는 pdfplumber) 필요", file=sys.stderr)
    sys.exit(2)

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl 필요", file=sys.stderr)
    sys.exit(2)


KOR_SEP = "ㅣ"  # 'ㅣ'
RE_YEAR = re.compile(r"(20\d{2})")
RE_NUM = re.compile(r"-?\d+(?:\.\d+)?")


def parse_meta(rel_parts, stem):
    univ = None
    if KOR_SEP in stem:
        head = stem.split(KOR_SEP, 1)[0].strip()
        if head:
            univ = head
    if univ is None:
        for comp in reversed(rel_parts[:-1]):
            m = re.search(r"([가-힣A-Za-z]+(?:대학교|대))", comp)
            if m:
                univ = m.group(1)
                break
    if univ is None:
        first = re.split(r"\s|_|-", stem)[0] if stem else ""
        if first.endswith("대") or first.endswith("교"):
            univ = first

    yr_m = RE_YEAR.search(stem)
    year = int(yr_m.group(1)) if yr_m else None

    if "수시" in stem:
        track = "susi"
    elif "정시" in stem:
        track = "jeongsi"
    else:
        track = "unknown"

    sel = []
    for k in ("학생부교과", "교과전형", "교과형"):
        if k in stem:
            sel.append("학생부교과")
            break
    for k in ("학생부종합", "종합전형", "종합형"):
        if k in stem:
            sel.append("학생부종합")
            break
    if "논술" in stem:
        sel.append("논술")
    if "실기" in stem or "예체능" in stem:
        sel.append("실기/예체능")
    if track == "jeongsi" and not sel:
        sel.append("정시")
    return {"univ": univ, "year": year, "track": track, "selection_types": sel}


def heuristic_rows_from_text(text):
    rows = []
    warns = []
    for raw in text.splitlines():
        line = raw.strip()
        if len(line) < 4:
            continue
        if not ('가' <= line[0] <= '힣'):
            continue
        nums_str = RE_NUM.findall(line)
        nums = [float(s) for s in nums_str]
        if len(nums) < 2:
            continue
        first_num_pos = RE_NUM.search(line).start()
        dept = line[:first_num_pos].strip()
        if not dept or len(dept) > 30:
            continue
        applicants = None
        comp = None
        grades = []
        for n in nums:
            if 1.0 <= n <= 9.5:
                grades.append(n)
            elif n.is_integer() and 10 <= n < 100000 and applicants is None:
                applicants = int(n)
            elif comp is None and 0.5 <= n < 1000 and not (1.0 <= n <= 9.5):
                comp = n
        rows.append({
            "department": dept,
            "selection": None,
            "applicants": applicants,
            "competition_rate": comp,
            "grade_50pct": grades[0] if grades else None,
            "grade_70pct": grades[1] if len(grades) > 1 else None,
            "grade_avg": grades[2] if len(grades) > 2 else None,
            "grade_min": grades[-1] if grades else None,
            "raw_cells": [line],
            "extraction_confidence": _confidence(nums, grades),
        })
    if not rows:
        warns.append("표 추출 실패: 모집단위 행을 인식하지 못했습니다.")
    return rows, warns


def _confidence(nums, grades):
    s = 0.0
    if len(nums) >= 4:
        s += 0.4
    elif len(nums) >= 2:
        s += 0.2
    if len(grades) >= 2:
        s += 0.3
    if grades and all(1.0 <= g <= 7.0 for g in grades):
        s += 0.2
    return min(1.0, s)


def extract_pdf(path):
    warns = []
    full = ""
    page_count = 0
    if _HAS_FITZ:
        try:
            with fitz.open(path) as doc:
                page_count = doc.page_count
                full = "\n".join(doc[i].get_text("text") for i in range(page_count))
        except Exception as e:
            warns.append(f"PyMuPDF 실패: {e}")
            full = ""
    if not full and pdfplumber is not None:
        try:
            with pdfplumber.open(path) as pdf:
                page_count = len(pdf.pages)
                full = "\n".join((p.extract_text() or "") for p in pdf.pages)
        except Exception as e:
            return [], [f"PDF 처리 오류: {e}"], {}
    if not full.strip():
        return [], ["PDF가 이미지/스캔본으로 보입니다 (텍스트 0)."], {"pages": page_count}
    rows, w = heuristic_rows_from_text(full)
    warns.extend(w)
    return rows, warns, {"pages": page_count, "chars": len(full)}


def extract_xlsx(path):
    warns = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return [], [f"XLSX 열기 실패: {e}"], {}
    rows = []
    sheet_count = 0
    for sn in wb.sheetnames:
        sheet_count += 1
        ws = wb[sn]
        prev_first = None
        for r in ws.iter_rows(values_only=True):
            cells = [("" if v is None else str(v)).strip() for v in r]
            if all(c == "" for c in cells):
                continue
            if cells and cells[0] == "" and prev_first:
                cells[0] = prev_first
            if cells and cells[0]:
                prev_first = cells[0]
            first = next((c for c in cells if c), "")
            if not first or not ('가' <= first[0] <= '힣'):
                continue
            nums = []
            for c in cells:
                cs = c.replace(",", "")
                try:
                    nums.append(float(cs))
                except Exception:
                    pass
            if len(nums) < 2:
                continue
            grades = sorted([n for n in nums if 1.0 <= n <= 9.5])
            comp = next((n for n in nums if 0.5 <= n < 1000 and not n.is_integer() and not (1.0 <= n <= 9.5)), None)
            applicants = next((int(n) for n in nums if n.is_integer() and 10 <= n < 100000), None)
            rows.append({
                "department": first,
                "selection": cells[1] if len(cells) > 1 and cells[1] else None,
                "applicants": applicants,
                "competition_rate": comp,
                "grade_50pct": grades[0] if grades else None,
                "grade_70pct": grades[1] if len(grades) > 1 else None,
                "grade_avg": grades[2] if len(grades) > 2 else None,
                "grade_min": grades[-1] if grades else None,
                "raw_cells": cells,
                "extraction_confidence": 0.7 if len(nums) >= 4 else 0.4,
            })
    wb.close()
    if not rows:
        warns.append("XLSX 표 추출 실패: 데이터 행을 인식하지 못했습니다.")
    return rows, warns, {"sheets": sheet_count}


def normalize(rows):
    out = []
    for r in rows:
        if not r["department"]:
            continue
        r["department"] = r["department"].strip()
        for k in ("grade_50pct", "grade_70pct", "grade_avg", "grade_min"):
            if r[k] is not None and not (1.0 <= r[k] <= 9.5):
                r[k] = None
        out.append(r)
    return out


def slug(univ, year, track):
    basis = f"{univ or '?'}|{year or 0}|{track}".encode()
    digest = hashlib.sha256(basis).hexdigest()[:8]
    return f"{track}-{year or 0}-{digest}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("data_dir", type=Path)
    ap.add_argument("out", type=Path)
    args = ap.parse_args()

    raw_dir = args.data_dir / "raw"
    docs = []
    t0 = time.time()
    for path in sorted(raw_dir.rglob("*")):
        if not path.is_file():
            continue
        ext = path.suffix.lower().lstrip(".")
        if ext not in {"pdf", "xlsx", "xls", "hwp", "hwpx", "jpg", "jpeg", "png", "html", "htm"}:
            continue
        rel = path.relative_to(raw_dir)
        rel_parts = list(rel.parts)
        if any(p.endswith("_files") for p in rel_parts[:-1]):
            continue
        stem = path.stem
        meta = parse_meta(rel_parts, stem)
        region = rel_parts[0] if rel_parts else "기타"
        size = path.stat().st_size
        rows, warns, info = [], [], {}
        if ext == "pdf":
            rows, warns, info = extract_pdf(path)
        elif ext in ("xlsx", "xls"):
            rows, warns, info = extract_xlsx(path)
        elif ext in ("hwp", "hwpx"):
            warns.append("HWP는 사전 변환 후 PDF로 처리됩니다.")
        elif ext in ("jpg", "jpeg", "png"):
            warns.append("이미지 자료. OCR 결과 머지는 별도 단계.")
        elif ext in ("html", "htm"):
            warns.append("HTML 자료는 본 추출기에서 미지원.")
        rows = normalize(rows)
        univ = meta["univ"] or "(미식별)"
        doc = {
            "id": slug(univ, meta["year"], meta["track"]),
            "univ": univ,
            "region": region,
            "year": meta["year"],
            "track": meta["track"],
            "selection_types": meta["selection_types"],
            "source": {
                "filename": path.name,
                "format": ext,
                "url": f"data/raw/{rel.as_posix()}",
                "size_bytes": size,
                "official_url": None,
                "license_status": "pending",
                "note": None,
            },
            "rows": rows,
            "ocr_text": None,
            "warnings": warns,
            "_extract_info": info,
        }
        docs.append(doc)
        sys.stderr.write(". " + rel.as_posix() + " -> " + str(len(rows)) + " rows\n")
    elapsed = time.time() - t0

    by_region = {}
    by_format = {}
    total_rows = 0
    conf_sum = 0.0
    conf_cnt = 0
    for d in docs:
        by_region[d["region"]] = by_region.get(d["region"], 0) + 1
        by_format[d["source"]["format"]] = by_format.get(d["source"]["format"], 0) + 1
        total_rows += len(d["rows"])
        for r in d["rows"]:
            conf_sum += r["extraction_confidence"]
            conf_cnt += 1
    summary = {
        "total_documents": len(docs),
        "total_rows": total_rows,
        "avg_confidence": (conf_sum / conf_cnt) if conf_cnt else 0.0,
        "by_region": dict(sorted(by_region.items())),
        "by_format": dict(sorted(by_format.items())),
        "elapsed_sec": round(elapsed, 2),
    }
    out = {
        "schema_version": 1,
        "generated_at": "unix:" + str(int(time.time())),
        "summary": summary,
        "documents": docs,
    }
    sys.stderr.write(f"[debug] writing to {args.out}\n"); sys.stderr.flush()
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    sys.stderr.write(f"[debug] wrote {args.out.stat().st_size} bytes\n"); sys.stderr.flush()

    avg_conf = summary["avg_confidence"]
    by_r = summary["by_region"]
    by_f = summary["by_format"]
    print()
    print(f"=== complete: {len(docs)} docs, {total_rows} rows, avg_conf {avg_conf:.2f}, {elapsed:.1f}s")
    print("by_region:", by_r)
    print("by_format:", by_f)


if __name__ == "__main__":
    main()

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    sys.stderr.write("[done] wrote " + str(args.out) + "\n")

    avg_conf = summary["avg_confidence"]
    print()
    print("complete:", len(docs), "docs,", total_rows, "rows, avg_conf %.3f, %.1fs" % (avg_conf, elapsed))
    print("by_region:", summary["by_region"])
    print("by_format:", summary["by_format"])


if __name__ == "__main__":
    main()
